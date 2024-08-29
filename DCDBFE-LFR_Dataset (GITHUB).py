
"""

# -*- coding: utf-8 -*-
"""
Created on Fri Aug 30 05:32:21 2024
read all files
@author: Siti Haryanti
"""

import networkx as nx
from sklearn.metrics.cluster import normalized_mutual_info_score
from sklearn.metrics.cluster import adjusted_rand_score
from collections import defaultdict
import time
import datetime
from sklearn import metrics
import math
import matplotlib.pyplot as plt  
import numpy as np
from numpy import linalg as LA
import openpyxl

start_time = time.time()

# global comm_list
# comm_list=[]


def str_to_int(x):
    return [[int(v) for v in line.split()] for line in x]


def node_addition(G, addnodes, communitys):  
    change_comm = set()  
    processed_edges = set()  

    for u in addnodes:
        neighbors_u = G.neighbors(u)
        neig_comm = set()  
        pc = set()
        for v in neighbors_u:
            if v in communitys:
                neig_comm.add(communitys[v])
            pc.add((u, v))
            pc.add((v, u))  
        if len(neig_comm) > 1:  
            change_comm = change_comm | neig_comm
            lab = max(communitys.values()) + 1
            communitys.setdefault(u, lab)  
            change_comm.add(lab)
        else:
            if len(neig_comm) == 1:  
                communitys.setdefault(u, list(neig_comm)[0])  
                processed_edges = processed_edges | pc
            else:
                communitys.setdefault(u, max(communitys.values()) + 1)  

    return change_comm, processed_edges, communitys  


def node_deletion(G, delnodes, communitys):  # tested, correct
    change_comm = set()  
    processed_edges = set()  
    for u in delnodes:
        neighbors_u = G.neighbors(u)
        neig_comm = set()  # 邻居所在社区标签
        for v in neighbors_u:
            if v in communitys:
                neig_comm.add(communitys[v])
            processed_edges.add((u, v))
            processed_edges.add((v, u))
        del communitys[u]  
        change_comm = change_comm | neig_comm
    return change_comm, processed_edges, communitys  


def edge_addition(addedges, communitys):  
    change_comm = set()  
    #    print addedges
    #    print communitys
    for item in addedges:
        neig_comm = set()  
        neig_comm.add(communitys[item[0]])  
        neig_comm.add(communitys[item[1]])
        if len(neig_comm) > 1:  
            change_comm = change_comm | neig_comm
    return change_comm  


def edge_deletion(deledges, communitys):  
    change_comm = set()  # 存放结构可能发现改变的社区标签
    for item in deledges:
        neig_comm = set()  # 邻居所在社区标签
        neig_comm.add(communitys[item[0]])  # 判断一边两端的节点所在社区
        neig_comm.add(communitys[item[1]])
        if len(neig_comm) == 1:  
            change_comm = change_comm | neig_comm

    return change_comm  

def getchangegraph(all_change_comm, newcomm, Gt):
    Gte = nx.Graph()
    com_key = newcomm.keys()
    for v in Gt.nodes():
        if v not in com_key or newcomm[v] in all_change_comm:
            Gte.add_node(v)
            neig_v = Gt.neighbors(v)
            for u in neig_v:
                if u not in com_key or newcomm[u] in all_change_comm:
                    Gte.add_edge(v, u)
                    Gte.add_node(u)

    return Gte


# CDBFE ****************************************************************

nodecount_comm = defaultdict(int)  


def CDBFE(G,algorithm_mode=3):
    deg = G.degree()

    # def AA(NA, NB):
    def AA(u, v):

        # comm_nodes = list(NA & NB)
        comm_nodes = list(u & v)
        #print("ini adalah node :", comm_nodes)
        sim = 0
        for node in comm_nodes:
            degnode = deg[node]
            if deg[node] == 1:
                degnode = 1.1
            # sim = sim + (1.0 / math.log(degnode))  ##AA similarity asal
            # sim = sum (1.0 / math.log(degnode))  ##AA similarity x boleh guna 'sum' kata eko ; ini salah
            sim = sim + (1.0 / (degnode))       ##RA similarity (index lain)
            # sim = sim + ((1.0 / math.log(degnode)) + (1.0 / (degnode)))  ##AA+RA hybrid
            # sim = sim + ((1.0 / math.sqrt(degnode)) + (1.0 / (degnode)*(degnode)))  ##RA1 modified
            # sim = sim + (1.0 / (math.log(degnode))*(math.log(degnode)))
            #print("ini adalah sim : ", sim)
        return sim

        # Compute the jaccard similarity coefficient of two node
    def simjkd(u, v):
        set_v = v
        set_u = u
        # jac = len(set_v & set_u) * 1.0 / len(set_v | set_u) #index Jaccard
        jac = len(set_v & set_u) * 1.0 / (len(set_v)*len(set_u))  ###tukar index Leicht-Holme-Newman
        return jac

    def simsltn(u, v):
        set_v = v
        set_u = u
        salton = len(set_v & set_u) * 1.0 / math.sqrt(len(v) * len(u)) #index Salton
        # salton = len(set_v & set_u) * 1.0 / (len(v)+len(u))  ##tukar index Sorensen Similarity
        return salton

    #coefficient
    def simjkdcoff(u,v):
        set_v = set(G.neighbors(v))
        set_v.add(v)
        set_u = set(G.neighbors(u))
        set_u.add(u)
        jac = len(set_v & set_u) * 1.0 / len(set_v | set_u)  # index Jaccard
        #jac = len(set_v & set_u) * 1.0 / (len(set_v) * len(set_u))  ###Leicht-Holme-Newman
        return jac

    def simsltncoff(u, v):
        set_v = set(G.neighbors(v))
        set_v.add(v)
        set_u = set(G.neighbors(u))
        set_u.add(u)
        # salton = len(set_v & set_u) * 1.0 / math.sqrt(len(set_v) * len(set_u))  # index Salton
        salton = len(set_v & set_u) * 1.0 / (len(set_v) + len(set_u))  ##Sorensen Similarity
        return salton

    # Initialize communities of each node
   
    node_community = dict(zip(G.nodes(), G.nodes()))

    # Compute the SUBMODULES (EQUATION VERTEX ATTRACTION)
 
    st = {}  # storge the AA
    # compute the AA
    for node in G.nodes():
        Nv = sorted(G.neighbors(node))
        for u in Nv:
            Nu = G.neighbors(u)
            keys = str(node) + '_' + str(u)
            #st.setdefault(keys, simsltn(set(Nv), set(Nu)))
            #st.setdefault(keys, simjkd(set(Nv), set(Nu)))
            #st.setdefault(keys, AA(set(Nv), set(Nu)))
            if algorithm_mode == 1 :
                st.setdefault(keys, AA(set(Nv), set(Nu)))
            elif algorithm_mode == 2 :
                st.setdefault(keys, simjkd(set(Nv), set(Nu)))
            elif algorithm_mode == 3 :
                st.setdefault(keys, simsltn(set(Nv), set(Nu)))
            else:
                exit()
    if algorithm_mode == 1:
        print('AA index,done')
    elif algorithm_mode == 2:
        print('JACCARD index,done')
    elif algorithm_mode == 3:
        print('SALTION index,done')
    else:
        pass

    for node in G.nodes():
        # The degree of each node
        deg_node = deg[node]
        flag = True
        maxsimdeg = 0
        selected = node
        if deg_node == 1:
            # node_community[node] =  node_community[ G.neighbors(node)[0]]
            node_community[node] = node_community[list(G.neighbors(node))[0]]
        else:
            for neig in G.neighbors(node):
                deg_neig = deg[neig]
                if flag is True and deg_node <= deg_neig:
                    flag = False
                    break

            if flag is False:
                for neig in sorted(G.neighbors(node)):
                    deg_neig = deg[neig]
                    # Compute the Jaccard similarity coefficient
                    # nodesim =  simjkd(node, neig)
                    # Use the AAindex
                    #keys = str(node) + '_' + str(neig)
                    #nodesim = st[keys]
                    if algoritm_mode == 1:
                        keysaa = str(node) + '_' + str(neig)
                        nodesim = st[keysaa]
                    elif algorithm_mode == 2:
                        #keys = str(node) + '_' + str(neig)
                        nodesim = simjkdcoff(node, neig)
                    elif algorithm_mode == 3:
                        nodesim = simsltncoff(node, neig)
                    else:
                        pass

                    # Compute the node attraction
                    nodesimdeg = deg_neig * nodesim
                    if nodesimdeg > maxsimdeg:
                        selected = neig
                        maxsimdeg = nodesimdeg
                    node_community[node] = node_community[selected]

## Simulate the BIRD FLOCK EFFECT
    old_persum = -(2 ** 63 - 1)
    old_netw_per = -(2 ** 63 - 1)

    persum = old_persum + 1
    netw_per = old_netw_per + 0.1
    maxit = 5
    itern = 0

    print("loop begin:")
    while itern < maxit:
        itern += 1
        old_netw_per = netw_per
        old_persum = persum
        persum = 0
        for node in G.nodes():
            neiglist = sorted(G.neighbors(node))
            cur_p = per(G, node, node_community)  #
            nodeneig_comm = nodecount_comm.keys()
            cur_p_neig = 0

            for neig in neiglist:
                cur_p_neig += per(G, neig, node_community)
    
            try :
                for neig_comm in nodeneig_comm:

                    node_pre_comm = node_community[node]
                    new_p_neig = 0
                    if node_pre_comm != neig_comm:
                        try:
                            node_community[node] = neig_comm
                            new_p = per(G, node, node_community)

                            if cur_p <= new_p:

                                if cur_p == new_p:
                                    for newneig in neiglist:
                                        new_p_neig += per(G, newneig, node_community)
                                    if cur_p_neig < new_p_neig:
                                        cur_p = new_p
                                        cur_p_neig = new_p_neig
                                    else:
                                        node_community[node] = node_pre_comm

                                else:
                                    for newneig in neiglist:
                                        new_p_neig += per(G, newneig, node_community)
                                    cur_p = new_p
                                    cur_p_neig = new_p_neig
                            else:
                                node_community[node] = node_pre_comm
                        except:
                            pass
            except:
                print('terjadi error')
            persum += cur_p
 
    print("loop done")
    return node_community


# The internal degree of node v in a community
# EQUATION MODULE ATTRACTIVENESS (MA)

def per(G, v, node_community):
    neiglist1 = G.neighbors(v)  # First-layer neighbors
    in_v = 0
    second_layer_in_v = 0  # Second-layer connectivity
    third_layer_in_v = 0  # Third-layer connectivity
    global nodecount_comm

    # First-layer connectivity
    for neig in neiglist1:
        if node_community[neig] == node_community[v]:
            in_v += 1
        else:
            nodecount_comm[node_community[neig]] += 1

    # Second-layer connectivity
    for neig in neiglist1:
        neiglist2 = G.neighbors(neig)  # Neighbors of neighbors (second layer)
        for neig2 in neiglist2:
            if neig2 != v and node_community[neig2] == node_community[v]:
                second_layer_in_v += 1

            # Third-layer connectivity
            neiglist3 = G.neighbors(neig2)  # Neighbors of neighbors of neighbors (third layer)
            for neig3 in neiglist3:
                if neig3 != neig and node_community[neig3] == node_community[v]:
                    third_layer_in_v += 1

    cin_v = 1.0 * (in_v * in_v) + second_layer_in_v + third_layer_in_v  # Include third-layer connectivity
    per = cin_v
    return per




# ****************************************************************************

def Errorrate(clusters, classes, n):

    A = np.zeros((n, len(clusters)), int)
    C = np.zeros((n, len(classes)), int)
    k = 0
    for nodelist in clusters:
        for node in nodelist:
            A[node - 1][k] = 1
        k = k + 1
    k = 0
    for nodelist in classes:
        for node in nodelist:
            C[node - 1][k] = 1
        k = k + 1
    t = A.dot(A.T) - C.dot(C.T)
    errors = LA.norm(t)
    return errors



def conver_comm_to_lab(comm1):  
    overl_community = {}
    for node_v, com_lab in comm1.items():
        if com_lab in overl_community.keys():
            overl_community[com_lab].append(node_v)
        else:
            overl_community.update({com_lab: [node_v]})
    return overl_community


def getscore(comm_true, comm_dete, num):
    actual = []
    baseline = []
    for j in range(len(comm_true)): 
        for c in comm_true[j]:  
            flag = False
            for k in range(len(comm_dete)):  
                if c in comm_dete[k] and flag == False:
                    flag = True
                    actual.append(j)
                    baseline.append(k)
                    break

    NMI1 = metrics.normalized_mutual_info_score(actual, baseline)
    ARI1 = metrics.adjusted_rand_score(actual, baseline)
    Purity1 = purity_score(baseline, actual)
    # errors=Errorrate(comm_dete,comm_true,num)
    errors = 0
    print('nmi', NMI1)
    print('ari', ARI1)
    print('rate error', errors)

    return NMI1, ARI1, errors


def drawcommunity(g, partition, filepath):
    pos = nx.spring_layout(g)
    count1 = 0
    t = 0
    node_color = ['#66CCCC', '#FFCC00', '#99CC33', '#CC6600', '#CCCC66', '#FF99CC', '#66FFFF', '#66CC66', '#CCFFFF',
                  '#CCCC00', '#CC99CC', '#FFFFCC']
    #    print(node_color[1])

    for com in set(partition.values()):
        count1 = count1 + 1.
        list_nodes = [nodes for nodes in partition.keys()
                      if partition[nodes] == com]

        nx.draw_networkx_nodes(g, pos, list_nodes, node_size=220,
                               node_color=node_color[t])
        nx.draw_networkx_labels(g, pos)
        t = t + 1

    nx.draw_networkx_edges(g, pos, with_labels=True, alpha=0.5)
    plt.savefig(filepath)
    plt.show()


############################################################
# ----------main-----------------
edges_added = set()
edges_removed = set()
nodes_added = set()
nodes_removed = set()
G = nx.Graph()
print("1. AA Similarity\n"
      "2. Jaccard Similarity\n"
      "3. Salton Similarity\n")
algoritm_similarity = input("Input name of Similarity Algorithm :")
algoritm_mode = None
if int(algoritm_similarity) == 1:
    algoritm_mode = 1
elif int(algoritm_similarity) == 2:
    algoritm_mode = 2
elif int(algoritm_similarity) == 3:
    algoritm_mode = 3
else:
    exit()

#allpath = './data/trydataset/files2.txt'
# allpath = './input/files.txt'
allpath = './input/merge40.txt'
with open(allpath, 'r') as f:
    allpathlist = f.readlines()
    f.close()
# allpathlists=allpathlist[0].strip('\n')
pathfile = ''
for pt in allpathlist:
    pathfile = pt.strip('\n')
    print(pathfile)
#    path = './data/trydataset/' + pathfile + '/'
    path = './input/' + pathfile + '/'
    edge_file = ''
    comm_file = ''
    G.clear()
    # read edgefile list, where storage the filename of each snapshot
    edgefilelist = []
    commfilelist = []
    with open(path + 'edgeslist.txt', 'r') as f:
        edgefilelist = f.readlines()
        f.close()
    edge_file = edgefilelist[0].strip('\n')
    with open(path + 'commlist.txt', 'r') as f:
        commfilelist = f.readlines()
        f.close()
    comm_file = commfilelist[0].strip('\n')

    # path='./LFR/t/'
    # path='./data/test/'
    with open(path + edge_file, 'r') as f:

        edge_list = f.readlines()
        for edge in edge_list:
            edge = edge.split()
            G.add_node(int(edge[0]))
            G.add_node(int(edge[1]))
            G.add_edge(int(edge[0]), int(edge[1]))
        f.close()
    G = G.to_undirected()

    nodenumber = G.number_of_nodes()
    with open(path + comm_file, 'r') as f:
        comm_list = f.readlines()
        comm_list = str_to_int(comm_list)  # 真实社区
        f.close()

  

    comm = {}  
    comm = CDBFE(G,algoritm_mode)  
    initcomm = conver_comm_to_lab(comm)
    comm_va = list(initcomm.values())
    commu_num = len(comm_va)
    tru_num = len(comm_list)
    NMI, ARI, Errors = getscore(comm_list, comm_va, nodenumber)
    import xlsxwriter, os
    if int(algoritm_similarity) == 1:
        file_exists = os.path.exists('result_score_LFR_aa.xlsx')
        if file_exists == True:
            path_score = 'result_score_LFR_aa.xlsx'
        else:
            workbook = xlsxwriter.Workbook('result_score_LFR_aa.xlsx')
            path_score = 'result_score_LFR_aa.xlsx'
            f = open(path_score,'a+')      
            f.write('path'+"\t"+'NMI'+"\t"+'ARI'+"\t"+'Purity'+'\t'+'detected_community_number'+'ture_community_number'+'errors'"\n" )
            #f.write(path+'_1'+"\t"+str(NMI)+"\t"+str(ARI)+"\t"+str(Purity)+'\t'+str(commu_num)+'\t'+str(tru_num)+str(Errors)+"\n" )
            f.close()
            wb=openpyxl.Workbook(path_score)
            wb.save(path_score)

    elif int(algoritm_similarity) == 2:

        file_exists = os.path.exists('result_score_LFR_jaccard.xlsx')
        if file_exists == True:
            path_score = 'result_score_LFR_jaccard.xlsx'
        else:
            workbook = xlsxwriter.Workbook('result_score_LFR_jaccard.xlsx')
            path_score = 'result_score_LFR_jaccard.xlsx'
            f = open(path_score,'a+')        #写入文件
            f.write('path'+"\t"+'NMI'+"\t"+'ARI'+"\t"+'Purity'+'\t'+'detected_community_number'+'ture_community_number'+'errors'"\n" )
            #f.write(path+'_1'+"\t"+str(NMI)+"\t"+str(ARI)+"\t"+str(Purity)+'\t'+str(commu_num)+'\t'+str(tru_num)+str(Errors)+"\n" )
            f.close()
            wb=openpyxl.Workbook(path_score)
            wb.save(path_score)

    elif int(algoritm_similarity) == 3:

        file_exists = os.path.exists('result_score_LFR_salton.xlsx')
        if file_exists == True:
            path_score = 'result_score_LFR_salton.xlsx'
        else:
            workbook = xlsxwriter.Workbook('result_score_LFR_salton.xlsx')
            path_score = 'result_score_LFR_salton.xlsx'
            f = open(path_score,'a+')        #写入文件
            f.write('path'+"\t"+'NMI'+"\t"+'ARI'\t'+'detected_community_number'+'true_community_number'+'errors'"\n" )
            #f.write(path+'_1'+"\t"+str(NMI)+"\t"+str(ARI)+"\t"+'\t'+str(commu_num)+'\t'+str(tru_num)+str(Errors)+"\n" )
            f.close()
            wb=openpyxl.Workbook(path_score)
            wb.save(path_score)
    else:
        exit()
 
    wb = openpyxl.load_workbook(filename=path_score)
    sheetname = path[:len(path) - 1].split('/')
    fix_pat = sheetname[len(sheetname)-1]
    print(fix_pat)
    #ws = wb.create_sheet(path[18:len(path) - 1])
    ws = wb.create_sheet(fix_pat)

    row = ['path', 'NMI', 'ARI',  'detected_community_number', 'true_community_number', 'errors']
    ws.append(row)
    row = ['1', str(NMI), str(ARI), str(commu_num), str(tru_num), str(Errors)]
    ws.append(row)

    start = time.time()
    G1 = nx.Graph()
    G2 = nx.Graph()
    G1 = G

    l = len(edgefilelist)
    for i in range(1, l):
        print('begin loop:', i)
        comm_new_file = open(path + commfilelist[i].strip('\n'), 'r')
        comm_new = comm_new_file.readlines()
        comm_new_file.close()
        comm_new = str_to_int(comm_new)

        edge_list_new_file = open(path + edgefilelist[i].strip('\n'), 'r')
        edge_list_new = edge_list_new_file.readlines()
        edge_list_new_file.close()

        for line in edge_list_new:
            temp = line.strip().split()
            G2.add_edge(int(temp[0]), int(temp[1]))

     
        total_nodes = set(G1.nodes()) | set(G2.nodes())

        nodes_added = set(G2.nodes()) - set(G1.nodes())
      
        nodes_removed = set(G1.nodes()) - set(G2.nodes())

        edges_added = set(G2.edges()) - set(G1.edges())
   
        edges_removed = set(G1.edges()) - set(G2.edges())

        all_change_comm = set()
 
        addn_ch_comm, addn_pro_edges, addn_commu = node_addition(G2, nodes_added, comm)

        edges_added = edges_added - addn_pro_edges  # 去掉已处理的边

        all_change_comm = all_change_comm | addn_ch_comm

        deln_ch_comm, deln_pro_edges, deln_commu = node_deletion(G1, nodes_removed, addn_commu)
        all_change_comm = all_change_comm | deln_ch_comm
        edges_removed = edges_removed - deln_pro_edges

        adde_ch_comm = edge_addition(edges_added, deln_commu)
        all_change_comm = all_change_comm | adde_ch_comm

        dele_ch_comm = edge_deletion(edges_removed, deln_commu)
        all_change_comm = all_change_comm | dele_ch_comm
 
        unchangecomm = () 
        newcomm = {}  
        newcomm = deln_commu  
        unchangecomm = set(newcomm.values()) - all_change_comm
        unchcommunity = {key: value for key, value in newcomm.items() if value in unchangecomm}  
        Gtemp = nx.Graph()
        Gtemp = getchangegraph(all_change_comm, newcomm, G2)

        unchagecom_maxlabe = 0
        if len(unchangecomm) > 0:
            unchagecom_maxlabe = max(unchangecomm)
        #    print('subG',Gtemp.edges())
        if Gtemp.number_of_edges() < 1:  
            comm = newcomm
        else:
            try:
                getnewcomm = CDME(Gtemp,algoritm_mode)
                print('===========')
                
                mergecomm = {}  # 合并字黄格式为｛节点：社区｝
                mergecomm.update(unchcommunity)
                mergecomm.update(getnewcomm)
                
                comm = mergecomm  
                detectcom = list(conver_comm_to_lab(comm).values())
                commu_num = len(detectcom)
                tru_num = len(comm_new)
            except:
                pass
            print (detectcom)
        nodenumber = G2.number_of_nodes()
       
        NMI, ARI, Errors = getscore(comm_new, detectcom, nodenumber)  

        G1.clear()
        G1.add_edges_from(G2.edges())
        G2.clear()
       
        row = [str(i + 1), str(NMI), str(ARI), str(commu_num), str(tru_num), str(Errors)]
        ws.append(row)
    wb.save(path_score)
print('all done')

print(time.time()-start, "second")

